from flask import Flask, jsonify, request
from flask_cors import CORS
import pdfplumber
import re
import os
from datetime import datetime, date
from openpyxl import load_workbook
from rapidfuzz import fuzz

app = Flask(__name__)
CORS(app)

# ---------------------------------------------------------------------------
# Conciliação (Comprovante x ERP)
# ---------------------------------------------------------------------------

def valor_float(v):
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    v = re.sub(r"[^\d,.\-]", "", str(v))
    if "," in v and "." in v:
        v = v.replace(".", "").replace(",", ".")
    elif "," in v:
        v = v.replace(",", ".")
    try:
        return round(float(v), 2)
    except ValueError:
        return None


def _so_digitos_mask(v):
    return re.sub(r"[^\d*]", "", v or "")


def doc_compativel_global(a, b):
    a, b = _so_digitos_mask(a), _so_digitos_mask(b)
    if not a or not b or len(a) != len(b):
        return False
    return all(x == "*" or y == "*" or x == y for x, y in zip(a, b))


def conciliar(comprovante, erp):
    """Compara favorecido, cpf/cnpj (suporta mascara com *), pagamento e valor."""

    def data_so_dia(v):
        return (v or "")[:10]

    nome_score = fuzz.token_sort_ratio(
        (comprovante.get("favorecido") or "").upper(),
        (erp.get("favorecido") or "").upper(),
    )

    doc_ok = doc_compativel_global(
        comprovante.get("cnpj_cpf"), erp.get("cpf") or erp.get("cnpj")
    )
    valor_ok = valor_float(comprovante.get("valor")) == valor_float(erp.get("valor"))
    data_ok = data_so_dia(comprovante.get("pagamento")) == data_so_dia(erp.get("pagamento"))

    return {
        "conciliado": doc_ok and valor_ok and data_ok,
        "score_nome": nome_score,
        "doc_ok": doc_ok,
        "valor_ok": valor_ok,
        "data_ok": data_ok,
    }


def tentar_match_agrupado(comprovante, erps, erp_usado=None):
    """Agrupa os ERPs pelo mesmo AP (ou, se não tiver AP, por cpf/cnpj+pagamento)
    e soma os valores. Se a soma de um grupo bater com o valor do comprovante
    e nome/doc/data também baterem, considera conciliado."""

    erp_usado = erp_usado or set()
    grupos = {}
    for j, erp in enumerate(erps):
        if j in erp_usado:
            continue
        chave = erp.get("AP")
        if chave is None:
            chave = (erp.get("cpf") or erp.get("cnpj"), erp.get("pagamento"))
        grupos.setdefault(chave, []).append(j)

    for chave, indices in grupos.items():
        if len(indices) < 2:
            continue

        soma = sum(valor_float(erps[idx].get("valor")) or 0 for idx in indices)
        primeiro = erps[indices[0]]

        nome_score = fuzz.token_sort_ratio(
            (comprovante.get("favorecido") or "").upper(),
            (primeiro.get("favorecido") or "").upper(),
        )
        doc_ok = doc_compativel_global(
            comprovante.get("cnpj_cpf"), primeiro.get("cpf") or primeiro.get("cnpj")
        )
        data_ok = (comprovante.get("pagamento") or "")[:10] == (primeiro.get("pagamento") or "")[:10]
        valor_ok = valor_float(comprovante.get("valor")) == round(soma, 2)

        if doc_ok and valor_ok and data_ok:
            return {
                "erps": indices,
                "soma_valor": round(soma, 2),
                "score_nome": nome_score,
            }

    return None


def conciliar_listas(comprovantes, erps):
    """Recebe a lista de comprovantes (PDF) e a lista do ERP (Excel) e devolve
    o resultado final: conciliados, não conciliados, e ERP sobrando."""

    if isinstance(comprovantes, dict):
        comprovantes = [comprovantes]
    if isinstance(erps, dict):
        erps = [erps]

    conciliados = []
    nao_conciliados = []
    erp_usado = set()

    for i, comprovante in enumerate(comprovantes):
        melhor = None
        for j, erp in enumerate(erps):
            if j in erp_usado:
                continue
            resultado = conciliar(comprovante, erp)
            if resultado["conciliado"]:
                melhor = (j, resultado)
                break
            if melhor is None or resultado["score_nome"] > melhor[1]["score_nome"]:
                melhor = (j, resultado)

        if melhor and melhor[1]["conciliado"]:
            j, resultado = melhor
            erp_match = erps[j]
            erp_usado.add(j)
            conciliados.append({
                "comprovante": i,
                "erp": j,
                "dados": {
                    "favorecido": comprovante.get("favorecido"),
                    "cnpj_cpf": comprovante.get("cnpj_cpf"),
                    "pagamento": comprovante.get("pagamento"),
                    "valor": comprovante.get("valor"),
                    "banco": comprovante.get("banco"),
                    "identificador": comprovante.get("identificador"),
                    "id_transacao": comprovante.get("id_transacao"),
                },
                "erp_dados": erp_match,
            })
            continue

        agrupado = tentar_match_agrupado(comprovante, erps, erp_usado)
        if agrupado:
            erp_usado.update(agrupado["erps"])
            conciliados.append({
                "comprovante": i,
                "erps_agrupados": agrupado["erps"],
                "soma_valor": agrupado["soma_valor"],
                "dados": {
                    "favorecido": comprovante.get("favorecido"),
                    "cnpj_cpf": comprovante.get("cnpj_cpf"),
                    "pagamento": comprovante.get("pagamento"),
                    "valor": comprovante.get("valor"),
                    "banco": comprovante.get("banco"),
                    "identificador": comprovante.get("identificador"),
                    "id_transacao": comprovante.get("id_transacao"),
                },
                "erp_dados": [erps[idx] for idx in agrupado["erps"]],
            })
            continue

        j, resultado = melhor if melhor else (None, {})
        motivos = []
        if not resultado.get("doc_ok"):
            motivos.append("cpf/cnpj não bate")
        if not resultado.get("valor_ok"):
            motivos.append("valor não bate")
        if not resultado.get("data_ok"):
            motivos.append("data não bate")

        nao_conciliados.append({
            "motivos": motivos,
            "comprovante": {
                "favorecido": comprovante.get("favorecido"),
                "cnpj_cpf": comprovante.get("cnpj_cpf"),
                "pagamento": comprovante.get("pagamento"),
                "valor": comprovante.get("valor"),
                "banco": comprovante.get("banco"),
                "identificador": comprovante.get("identificador"),
                "id_transacao": comprovante.get("id_transacao"),
            },
        })

    erp_nao_conciliados = [erps[j] for j in range(len(erps)) if j not in erp_usado]

    return {
        "conciliados": conciliados,
        "nao_conciliados": nao_conciliados,
        "erp_nao_conciliados": erp_nao_conciliados,
    }

# ---------------------------------------------------------------------------
# Conversor XLSX (ERP) -> JSON
# ---------------------------------------------------------------------------

MAPEAMENTO_COLUNAS = {
    "PRO_IN_REDUZIDO": "UN",
    "AGN_ST_NOME": "favorecido",
    "AGN_ST_CGC": "cnpj",
    "AGN_ST_CPF": "cpf",
    "MOV_ST_DOCUMENTO": "Nº Documento",
    "MOV_REF_DT_DATADOCTO": "pagamento",
    "BAN_IN_NUMERO": "Banco",
    "CTA_ST_NUMERO": "Conta Bancaria",
    "LPR_RE_VALOR": "valor",
    "CHEQ_TPD_ST_CODIGO": "Forma de Pagamento",
    "CPA_IN_AP": "AP",
    "FPA_TPD_ST_CODIGO": "Tipo de Documento",
}

NOME_ABA = None  # ex: "Gd_Dados"


def tratar_valor_xlsx(valor):
    """Converte valores de data para string dd/mm/aaaa, remove espaços extras do texto."""
    if valor is None:
        return ""
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, str):
        texto = " ".join(valor.split())
        texto = texto.replace(" -", "-")
        return texto
    return valor


def tratar_numero_xlsx(valor):
    """Garante que o valor monetário vire float (numero), mesmo se vier como texto."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto:
        return None
    texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return valor


def converter_xlsx_para_json(caminho_xlsx):
    wb = load_workbook(filename=caminho_xlsx, data_only=True)
    aba = wb[NOME_ABA] if NOME_ABA else wb.active

    linhas = aba.iter_rows(values_only=True)
    cabecalho = next(linhas)

    indice_para_nome_saida = {}
    for indice, nome_coluna in enumerate(cabecalho):
        if nome_coluna in MAPEAMENTO_COLUNAS:
            indice_para_nome_saida[indice] = MAPEAMENTO_COLUNAS[nome_coluna]

    resultado = []
    for linha in linhas:
        if linha is None or all(c is None for c in linha):
            continue

        registro = {}
        for indice, nome_saida in indice_para_nome_saida.items():
            valor = linha[indice] if indice < len(linha) else None

            if nome_saida == "valor":
                valor = tratar_numero_xlsx(valor)
            else:
                valor = tratar_valor_xlsx(valor)

            registro[nome_saida] = valor

        cpf = str(registro.get("cpf") or "").strip()
        cnpj = str(registro.get("cnpj") or "").strip()
        if cpf:
            registro["cnpj"] = None
        elif cnpj:
            registro["cpf"] = None

        resultado.append(registro)

    return resultado

# ---------------------------------------------------------------------------
# Utilidades comuns
# ---------------------------------------------------------------------------

def extrair_texto_pdf(caminho_pdf):
    texto_total = ""
    with pdfplumber.open(caminho_pdf) as pdf:
        for page in pdf.pages:
            texto_pagina = page.extract_text(layout=True)
            if texto_pagina:
                texto_total += texto_pagina + "\n"
    return texto_total


def detectar_layout(texto_total):
    """Identifica qual banco/layout o PDF representa."""
    if "COMPROVANTE DE EFETIVAÇÃO DE PAGAMENTO PIX" in texto_total:
        return "sicoob"
    if "PAGAMENTO A FORNECEDORES" in texto_total:
        return "santander"
    return "desconhecido"


def normalizar_registro(banco_origem, favorecido=None, cnpj_cpf=None, banco=None,
                         agencia=None, conta_corrente=None, pagamento=None,
                         valor=None, identificador=None, id_transacao=None,
                         tipo_comprovante=None):
    """Garante um schema único de saída para qualquer layout."""
    def limpar(s):
        if s is None:
            return None
        s = re.sub(r'\s+', ' ', s).strip()
        return s or None

    return {
        "banco_origem": banco_origem,
        "tipo_comprovante": tipo_comprovante,
        "favorecido": limpar(favorecido),
        "cnpj_cpf": limpar(cnpj_cpf),
        "banco": limpar(banco),
        "agencia": limpar(agencia),
        "conta_corrente": limpar(conta_corrente),
        "pagamento": limpar(pagamento),
        "valor": limpar(valor),
        "identificador": limpar(identificador),
        "id_transacao": limpar(id_transacao),
    }


# ---------------------------------------------------------------------------
# Layout Sicoob (PIX)
# ---------------------------------------------------------------------------

def _sicoob_extrair_secao(bloco, inicio, fins):
    padrao = rf'{inicio}:\s*\n?(.*?)(?=' + '|'.join(fins) + r')'
    m = re.search(padrao, bloco, re.DOTALL)
    return m.group(1) if m else ""


def _sicoob_parsear_secao_destinatario(secao_texto):
    ordem = ["Nome", "CPF/CNPJ", "Instituição/Banco"]
    padrao_rotulo = re.compile(rf'^\s*({"|".join(ordem)}):\s*(.*)')

    campos = {c: "" for c in ordem}
    campo_atual = "Nome"

    for linha in secao_texto.split("\n"):
        linha_strip = linha.strip()
        if not linha_strip:
            continue

        m = padrao_rotulo.match(linha)
        if m:
            rotulo, conteudo = m.group(1), m.group(2).strip()
            if campos[rotulo] == "":
                campo_atual = rotulo
                campos[campo_atual] = conteudo
                if rotulo == "CPF/CNPJ":
                    campo_atual = "Instituição/Banco"
            elif conteudo:
                campos[campo_atual] = (campos[campo_atual] + " " + conteudo).strip()
        elif campo_atual:
            campos[campo_atual] = (campos[campo_atual] + " " + linha_strip).strip()

    return campos


def extrair_comprovantes_sicoob(texto_total):
    comprovantes = []
    blocos = re.split(r'COMPROVANTE DE EFETIVAÇÃO DE PAGAMENTO PIX', texto_total)[1:]

    for bloco in blocos:
        secao_destinatario = _sicoob_extrair_secao(bloco, "Destinatário", ["Dados do pagamento:"])
        campos_dest = _sicoob_parsear_secao_destinatario(secao_destinatario)

        nome = campos_dest["Nome"]
        cnpj = campos_dest["CPF/CNPJ"]
        banco = campos_dest["Instituição/Banco"]

        data_pagamento = re.search(r'Data do pagamento:\s*(\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2})', bloco)
        valor = re.search(r'Valor:\s*R\$\s*([\d.,]+)', bloco)
        identificador = re.search(r'Identificador:\s*(\S+)', bloco)
        id_transacao = re.search(r'ID Transação:\s*(\S+)', bloco)

        comprovantes.append(normalizar_registro(
            banco_origem="sicoob",
            tipo_comprovante="pix",
            favorecido=nome,
            cnpj_cpf=cnpj,
            banco=banco,
            pagamento=data_pagamento.group(1) if data_pagamento else None,
            valor=valor.group(1) if valor else None,
            identificador=identificador.group(1) if identificador else None,
            id_transacao=id_transacao.group(1) if id_transacao else None,
        ))

    return comprovantes


# ---------------------------------------------------------------------------
# Layout Santander (TED / Boleto)
# ---------------------------------------------------------------------------

def _santander_dividir_colunas(linha):
    """Divide uma linha em colunas, usando 2+ espaços como separador."""
    return [c.strip() for c in re.split(r'\s{2,}', linha.strip()) if c.strip()]


def _santander_eh_linha_separadora(linha):
    """Verifica se a linha é só uma linha de sublinhado (separador)."""
    return bool(re.fullmatch(r'_+\s*', linha.strip()))


def _santander_parsear_boleto(bloco):
    linhas = bloco.split("\n")

    favorecido = None
    cnpj = None
    banco = None

    for i, linha in enumerate(linhas):
        if "Nome/Razão Social do Beneficiário Original" in linha:
            for j in range(i + 1, min(i + 4, len(linhas))):
                if _santander_eh_linha_separadora(linhas[j]):
                    break
                cols = _santander_dividir_colunas(linhas[j])
                if cols:
                    favorecido = cols[0]
                    cnpj = cols[1] if len(cols) > 1 else None
                    break
        if "Instituição Financeira Favorecida" in linha:
            for j in range(i + 1, min(i + 3, len(linhas))):
                if _santander_eh_linha_separadora(linhas[j]):
                    break
                cols = _santander_dividir_colunas(linhas[j])
                if cols:
                    banco = cols[0]
                    break

    data_credito = re.search(r'Data do Crédito.*?\n.*?(\d{2}/\d{2}/\d{4})', bloco)
    valor = re.search(r'Valor a Pagar\s*\n.*?[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+([\d.,]+)', bloco)
    if not valor:
        valor = re.search(r'Valor Nominal.*?\n([\d.,]+)', bloco)
    autenticacao = re.search(r'Autenticação Bancária\s*\n\s*(\S+)', bloco)
    no_compromisso = re.search(r'No\. compromisso banco.*?\n\s*(\S+)', bloco)

    return normalizar_registro(
        banco_origem="santander",
        tipo_comprovante="boleto",
        favorecido=favorecido,
        cnpj_cpf=cnpj,
        banco=banco,
        pagamento=data_credito.group(1) if data_credito else None,
        valor=valor.group(1) if valor else None,
        identificador=no_compromisso.group(1) if no_compromisso else None,
        id_transacao=autenticacao.group(1) if autenticacao else None,
    )


def _santander_parsear_ted(bloco):
    linhas = bloco.split("\n")

    favorecido = None
    cnpj = None
    banco_ispb = None
    agencia = None
    conta_corrente = None
    valor_destinatario = None

    for i, linha in enumerate(linhas):
        if "Dados do Destinatário" in linha:
            for j in range(i + 1, min(i + 4, len(linhas))):
                if "Nome" in linhas[j] and "CNPJ/CPF" in linhas[j]:
                    if j + 1 < len(linhas) and not _santander_eh_linha_separadora(linhas[j + 1]):
                        cols = _santander_dividir_colunas(linhas[j + 1])
                        if cols:
                            favorecido = cols[0]
                            cnpj = cols[1] if len(cols) > 1 else None
                    break
        if "Banco/ISPB" in linha and "Agência" in linha:
            if i + 1 < len(linhas) and not _santander_eh_linha_separadora(linhas[i + 1]):
                cols = _santander_dividir_colunas(linhas[i + 1])
                if len(cols) >= 1:
                    banco_ispb = cols[0]
                if len(cols) >= 2:
                    agencia = cols[1]
                if len(cols) >= 3:
                    conta_corrente = cols[2]
                if len(cols) >= 4:
                    valor_destinatario = cols[3]

    data_credito = re.search(r'Data do Crédito.*?\n.*?(\d{2}/\d{2}/\d{4})', bloco)
    autenticacao = re.search(r'Autenticação Bancária\s*\n\s*(\S+)', bloco)
    no_compromisso = re.search(r'No\. compromisso banco.*?\n\s*(\S+)', bloco)

    return normalizar_registro(
        banco_origem="santander",
        tipo_comprovante="ted",
        favorecido=favorecido,
        cnpj_cpf=cnpj,
        banco=banco_ispb,
        agencia=agencia,
        conta_corrente=conta_corrente,
        pagamento=data_credito.group(1) if data_credito else None,
        valor=valor_destinatario,
        identificador=no_compromisso.group(1) if no_compromisso else None,
        id_transacao=autenticacao.group(1) if autenticacao else None,
    )


def extrair_comprovantes_santander(texto_total):
    comprovantes = []
    blocos = re.split(r'PAGAMENTO A FORNECEDORES', texto_total)[1:]

    for bloco in blocos:
        if "Comprovante de Crédito ao Favorecido" in bloco:
            comprovantes.append(_santander_parsear_ted(bloco))
        else:
            comprovantes.append(_santander_parsear_boleto(bloco))

    return comprovantes


# ---------------------------------------------------------------------------
# Orquestração
# ---------------------------------------------------------------------------

def extrair_comprovantes(caminho_pdf):
    """Detecta o layout do PDF e chama o parser apropriado."""
    texto_total = extrair_texto_pdf(caminho_pdf)
    layout = detectar_layout(texto_total)

    if layout == "sicoob":
        return extrair_comprovantes_sicoob(texto_total)
    elif layout == "santander":
        return extrair_comprovantes_santander(texto_total)
    else:
        return []


# ---------------------------------------------------------------------------
# Rotas Flask
# ---------------------------------------------------------------------------

@app.route("/", methods=["GET"])
def raiz():
    """Rota simples só para confirmar no navegador que o Flask está no ar."""
    return jsonify({"status": "ok", "servico": "finfortes-conciliacao"})


@app.route("/processar", methods=["POST"])
def processar():
    if 'arquivo' not in request.files:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400

    arquivo = request.files['arquivo']
    caminho_temp = os.path.join("/tmp", arquivo.filename)
    arquivo.save(caminho_temp)

    try:
        comprovantes = extrair_comprovantes(caminho_temp)
    finally:
        os.remove(caminho_temp)

    if not comprovantes:
        return jsonify({"erro": "Layout não reconhecido ou nenhum comprovante encontrado"}), 422

    return jsonify({"comprovantes": comprovantes})


@app.route("/processar_xlsx", methods=["POST"])
def processar_xlsx():
    if 'arquivo' not in request.files:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400

    arquivo = request.files['arquivo']
    caminho_temp = os.path.join("/tmp", arquivo.filename)
    arquivo.save(caminho_temp)

    try:
        dados = converter_xlsx_para_json(caminho_temp)
    except Exception as e:
        return jsonify({"erro": f"Erro ao processar planilha: {str(e)}"}), 422
    finally:
        os.remove(caminho_temp)

    return jsonify({"dados": dados})


@app.route("/conciliar", methods=["POST"])
def conciliar_rota():
    if 'pdf' not in request.files or 'xlsx' not in request.files:
        return jsonify({"erro": "Envie o PDF e a planilha juntos"}), 400

    arquivo_pdf = request.files['pdf']
    arquivo_xlsx = request.files['xlsx']

    caminho_pdf = os.path.join("/tmp", arquivo_pdf.filename)
    caminho_xlsx = os.path.join("/tmp", arquivo_xlsx.filename)
    arquivo_pdf.save(caminho_pdf)
    arquivo_xlsx.save(caminho_xlsx)

    try:
        comprovantes = extrair_comprovantes(caminho_pdf)
        erp = converter_xlsx_para_json(caminho_xlsx)
    except Exception as e:
        return jsonify({"erro": f"Erro ao processar arquivos: {str(e)}"}), 422
    finally:
        os.remove(caminho_pdf)
        os.remove(caminho_xlsx)

    if not comprovantes:
        return jsonify({"erro": "Layout do PDF não reconhecido ou nenhum comprovante encontrado"}), 422

    resultado = conciliar_listas(comprovantes, erp)
    resultado["log_pdf"] = comprovantes
    resultado["log_erp"] = erp

    return jsonify(resultado)


# Usado só se você rodar "python app.py" localmente pra testar.
# No PythonAnywhere, quem sobe o Flask é o arquivo WSGI (from app import app as application),
# então este bloco abaixo simplesmente não é executado lá — pode deixar como está.
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)